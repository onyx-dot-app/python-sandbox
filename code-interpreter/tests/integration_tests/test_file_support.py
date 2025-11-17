from __future__ import annotations

import io

from fastapi.testclient import TestClient

from app.main import create_app


def _client() -> TestClient:
    return TestClient(create_app())


def _upload_file(client: TestClient, filename: str, content: bytes) -> str:
    """Upload a file and return its file_id."""
    files = {"file": (filename, io.BytesIO(content), "application/octet-stream")}
    response = client.post("/v1/files", files=files)
    assert response.status_code == 201, f"Upload failed: {response.text}"
    file_id: str = response.json()["file_id"]
    return file_id


def _get_file(payload: dict[str, object], path: str) -> dict[str, object]:
    files = payload.get("files")
    assert isinstance(files, list), "response files must be a list"
    for entry in files:
        if isinstance(entry, dict) and entry.get("path") == path:
            return entry
    raise AssertionError(f"file '{path}' not found in response payload")


def test_execute_stages_provided_file() -> None:
    client = _client()

    # Upload the file first
    file_id = _upload_file(client, "data.txt", b"seeded content\n")

    response = client.post(
        "/v1/execute",
        json={
            "code": "from pathlib import Path\nprint(Path('inputs/data.txt').read_text())",
            "stdin": None,
            "timeout_ms": 1000,
            "files": [
                {
                    "path": "inputs/data.txt",
                    "file_id": file_id,
                }
            ],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["stdout"] == "seeded content\n\n"
    assert payload["stderr"] == ""
    # File was not modified, so it should NOT be in the output
    files = payload["files"]
    assert isinstance(files, list)
    assert len(files) == 0


def test_execute_returns_new_file_snapshot() -> None:
    client = _client()

    response = client.post(
        "/v1/execute",
        json={
            "code": (
                "from pathlib import Path\n"
                "Path('outputs').mkdir(parents=True, exist_ok=True)\n"
                "Path('outputs/result.txt').write_text('new file contents')"
            ),
            "stdin": None,
            "timeout_ms": 1000,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    created = _get_file(payload, "outputs/result.txt")
    assert created["kind"] == "file"
    # File should have a file_id
    file_id = created.get("file_id")
    assert isinstance(file_id, str)

    # Download and verify content
    download_response = client.get(f"/v1/files/{file_id}")
    assert download_response.status_code == 200
    assert download_response.content.decode("utf-8") == "new file contents"


def test_execute_can_edit_staged_file() -> None:
    client = _client()

    # Upload the file first
    file_id = _upload_file(client, "data.txt", b"initial")

    response = client.post(
        "/v1/execute",
        json={
            "code": (
                "from pathlib import Path\n"
                "path = Path('data.txt')\n"
                "path.write_text(path.read_text() + 'updated')"
            ),
            "stdin": None,
            "timeout_ms": 1000,
            "files": [
                {
                    "path": "data.txt",
                    "file_id": file_id,
                }
            ],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    updated_file = _get_file(payload, "data.txt")
    output_file_id = updated_file.get("file_id")
    assert isinstance(output_file_id, str)

    # Download and verify updated content
    download_response = client.get(f"/v1/files/{output_file_id}")
    assert download_response.status_code == 200
    assert download_response.content.decode("utf-8") == "initialupdated"


def test_execute_creates_file_and_returns_it() -> None:
    client = _client()

    response = client.post(
        "/v1/execute",
        json={
            "code": (
                "from pathlib import Path\n"
                "Path('new_file.txt').write_text('Hello, World!')\n"
                "print('File created')"
            ),
            "stdin": None,
            "timeout_ms": 1000,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["stdout"] == "File created\n"
    assert payload["stderr"] == ""

    # Verify the file is in the response
    created_file = _get_file(payload, "new_file.txt")
    assert created_file["kind"] == "file"
    file_id = created_file.get("file_id")
    assert isinstance(file_id, str)

    # Download and verify content
    download_response = client.get(f"/v1/files/{file_id}")
    assert download_response.status_code == 200
    assert download_response.content.decode("utf-8") == "Hello, World!"


def test_execute_rejects_illegal_path() -> None:
    client = _client()

    # Upload the file first
    file_id = _upload_file(client, "escape.txt", b"should fail")

    response = client.post(
        "/v1/execute",
        json={
            "code": "print('noop')",
            "stdin": None,
            "timeout_ms": 1000,
            "files": [
                {
                    "path": "../escape.txt",
                    "file_id": file_id,
                }
            ],
        },
    )

    assert response.status_code == 422
    payload = response.json()
    assert "File paths" in payload["detail"]


def test_execute_analyzes_excel_file() -> None:
    """Test that we can upload and analyze an Excel file using openpyxl."""
    import pathlib

    client = _client()

    # Read the sample Excel file from the test data directory
    test_data_dir = pathlib.Path(__file__).parent / "data"
    excel_file = test_data_dir / "Financial Sample.xlsx"
    excel_content = excel_file.read_bytes()

    # Upload the Excel file
    file_id = _upload_file(client, "financial_sample.xlsx", excel_content)

    # Analyze the Excel file - count rows, sum a column, and get some values
    response = client.post(
        "/v1/execute",
        json={
            "code": (
                "import openpyxl\n"
                "wb = openpyxl.load_workbook('financial_sample.xlsx')\n"
                "ws = wb.active\n"
                "# Get basic info\n"
                "row_count = ws.max_row\n"
                "col_count = ws.max_column\n"
                "# Get headers (first row)\n"
                "headers = [cell.value for cell in ws[1]]\n"
                "# Calculate sum of 'Units Sold' column (column H, index 8)\n"
                "units_sold_sum = sum(cell.value for cell in ws['H'][1:]"
                " if isinstance(cell.value, (int, float)))\n"
                "# Get first data row (row 2) country\n"
                "first_country = ws['A2'].value\n"
                "print(f'Rows: {row_count}')\n"
                "print(f'Columns: {col_count}')\n"
                "print(f'First column header: {headers[0]}')\n"
                "print(f'Units Sold Total: {units_sold_sum}')\n"
                "print(f'First Country: {first_country}')\n"
            ),
            "stdin": None,
            "timeout_ms": 5000,
            "files": [
                {
                    "path": "financial_sample.xlsx",
                    "file_id": file_id,
                }
            ],
        },
    )

    assert response.status_code == 200, f"Execution failed: {response.text}"
    payload = response.json()
    stdout = payload["stdout"]

    # Verify the analysis results
    assert "Rows: 701" in stdout
    assert "Columns: 16" in stdout
    assert "First column header: Segment" in stdout
    assert "Units Sold Total: 127931598.5" in stdout
    assert "First Country: Government" in stdout
